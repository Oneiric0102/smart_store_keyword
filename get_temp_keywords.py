import requests
import json
from bs4 import BeautifulSoup
import datetime as dt
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import os

headers = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36(KHTML, like Gecko) Chrome/86.0.4240.75 Safari/537.36"
    )
}

temp_keyword_list = []


# 네이버 검색 자동완성어 추출
def naver_search_keyword(coreKeyword):
    url = f"https://ac.search.naver.com/nx/ac?q={coreKeyword}&con=0&frm=nv&ans=2&r_format=json&r_enc=UTF-8&r_unicode=0&t_koreng=1&run=2&rev=4&q_enc=UTF-8&st=100&_callback=_jsonp_5"

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        json_string = response.text[9:-1]
        auto_keywords = json.loads(json_string)  # dict로 변환

        for auto_words in auto_keywords["items"]:
            for word in auto_words:
                temp_keyword_list.append(word[0].replace(" ", ""))
    except requests.exceptions.RequestException as e:
        print("검색 자동완성 Error occurred: ", e)


# 네이버 검색 연관검색어 추출
def naver_search_rel_keyword(coreKeyword):
    url = "https://search.naver.com/search.naver?"
    params = {"where": "web", "fbm": "1", "query": coreKeyword, "page": "2"}

    try:
        resp = requests.get(url, params)
        resp.raise_for_status()

        soup = BeautifulSoup(resp.text, "html.parser")
        shop_rel_words = soup.select("div.related_srch > ul > li")

        for i in shop_rel_words:
            temp_keyword_list.append(i.text.replace(" ", ""))
    except requests.exceptions.RequestException as e:
        print("검색 연관검색 Error occurred: ", e)


# 네이버 쇼핑 자동완성어 추출
def nshopping_keyword(coreKeyword):
    url = (
        "https://shopping.naver.com/api/modules/gnb/auto-complete?_vc_=1706983611197&keyword="
        + coreKeyword
    )

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        response = response.json()

        for auto_words in response["items"][1]:
            temp_keyword_list.append(auto_words[0][0].replace(" ", ""))
    except requests.exceptions.RequestException as e:
        print("쇼핑 자동완성 Error occurred: ", e)


# 네이버 쇼핑 연관검색어 추출
def nshopping_rel_keyword(coreKeyword):
    headers = {
        "Referer": (
            "https://search.shopping.naver.com/search/all?query="
            + coreKeyword.encode("utf-8").decode("iso-8859-1")
        ),
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36(KHTML, like Gecko) Chrome/86.0.4240.75 Safari/537.36"
        ),
        "Content-Type": ("application/x-www-form-urlencoded; charset=UTF-8"),
    }
    shop_rel_url = "https://search.shopping.naver.com/search/all?query=" + coreKeyword

    try:
        shop_rel_response = requests.get(shop_rel_url, headers=headers)
        shop_rel_response.raise_for_status()
        shop_rel_soup = BeautifulSoup(shop_rel_response.text, "html.parser")
        shop_rel_words = shop_rel_soup.select(
            "div.relatedTags_relation_srh__YG9s7 > ul > li"
        )
        if len(shop_rel_words):
            for li in shop_rel_words:
                temp_keyword_list.append(li.text.replace(" ", ""))
    except requests.exceptions.RequestException as e:
        print("쇼핑 연관검색 Error occurred: ", e)


# 네이버 쇼핑인사이트 TOP500 추출
def nshopping_insight_top500(cid):
    date = dt.datetime.now()
    startDate = (
        str(date.year - 1)
        + "-"
        + str(date.month).zfill(2)
        + "-"
        + str(date.day).zfill(2)
    )
    endDate = (
        str(date.year) + "-" + str(date.month).zfill(2) + "-" + str(date.day).zfill(2)
    )
    url = "https://datalab.naver.com/shoppingInsight/getCategoryKeywordRank.naver"
    headers = {
        "Referer": ("https://datalab.naver.com/shoppingInsight/sCategory.naver"),
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36(KHTML, like Gecko) Chrome/86.0.4240.75 Safari/537.36"
        ),
        "Content-Type": ("application/x-www-form-urlencoded; charset=UTF-8"),
    }

    for page in range(1, 26):
        data = {
            "cid": str(cid),
            "timeUnit": "date",
            "startDate": startDate,
            "endDate": endDate,
            "page": str(page),
            "count": "20",
            "age": "10,20,30,40,50,60",
            "gender": "f,m",
            "device": "pc,mo",
        }

        try:
            response = requests.post(url, headers=headers, data=data)
            response.raise_for_status()
            response = response.json()
            for keyword_info in response["ranks"]:
                temp_keyword_list.append(keyword_info["keyword"].replace(" ", ""))
            time.sleep(0.5)
        except requests.exceptions.RequestException as e:
            print("Error occurred: ", e)


# 추출한 키워드 후보들을 엑셀로 저장
def export_temp_keyword_list(coreKeyword):
    write_wb = Workbook()
    write_ws = write_wb.create_sheet("Sheet1")
    write_ws = write_wb.active

    for i in range(0, len(temp_keyword_list)):
        write_ws["A" + str(i + 1)] = temp_keyword_list[i]
    write_wb.save("./" + coreKeyword + "_temp.xlsx")


while True:
    temp_keyword_list = []
    search_keyword = input("메인 키워드 입력: ")
    search_cid = input("카테고리 cid 입력: ")

    temp_keyword_list.append(search_keyword)

    naver_search_keyword(search_keyword)
    naver_search_rel_keyword(search_keyword)
    nshopping_rel_keyword(search_keyword)
    nshopping_keyword(search_keyword)
    nshopping_insight_top500(search_cid)
    temp_keyword_list = list(set(temp_keyword_list))
    export_temp_keyword_list(search_keyword)
