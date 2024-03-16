import requests
import json
from bs4 import BeautifulSoup
import time
import signiturehelper
import os
from openpyxl import Workbook
from openpyxl import load_workbook

BASE_DIR = "./"
secret_file = os.path.join(BASE_DIR, "secrets.json")

with open(secret_file) as f:
    secrets = json.loads(f.read())


def get_secret(setting, secrets=secrets):
    try:
        return secrets[setting]
    except KeyError:
        err_msg = f"set the {setting} enviroment variable"
        raise print(err_msg)


ad_api_key = get_secret("AD_API_KEY")
ad_secret_key = get_secret("AD_SECRET_KEY")
ad_customer_id = get_secret("AD_CUSTOMER_ID")

naver_client_id = get_secret("NAVER_CLIENT_ID")
naver_client_secret = get_secret("NAVER_CLIENT_SECRET")

headers = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36(KHTML, like Gecko) Chrome/86.0.4240.75 Safari/537.36"
    )
}

keyword_base = {
    "keyword": "",
    "pcQcCnt": "",
    "mobileQcCnt": "",
    "totalQcCnt": "",
    "productsCnt": "",
    "ratio": "",
    "category": "",
}

result = {"keywords": []}
duplicate_check_list = []


# 검색광고 api 헤더 추출용
def get_header(method, uri, api_key, secret_key, customer_id):
    timestamp = str(round(time.time() * 1000))
    signature = signiturehelper.Signature.generate(timestamp, method, uri, secret_key)

    return {
        "Content-Type": "application/json; charset=UTF-8",
        "X-Timestamp": timestamp,
        "X-API-KEY": api_key,
        "X-Customer": str(customer_id),
        "X-Signature": signature,
    }


# 검색광고 api 사용
def naver_searchad_api(hintKeywords):
    BASE_URL = "https://api.naver.com"
    API_KEY = ad_api_key
    SECRET_KEY = ad_secret_key
    CUSTOMER_ID = ad_customer_id

    uri = "/keywordstool"
    method = "GET"

    params = {}

    params["hintKeywords"] = hintKeywords
    params["showDetail"] = "1"
    params["monthlyMobileQcCnt"] = ">1000"

    response = requests.get(
        BASE_URL + uri,
        params=params,
        headers=get_header(method, uri, API_KEY, SECRET_KEY, CUSTOMER_ID),
    ).json()

    keyword_list = response["keywordList"]

    api_result = keyword_list[0 : len(hintKeywords)]
    for i in range(len(hintKeywords) + 1, min(100, len(keyword_list))):
        keyword = keyword_list[i]
        if keyword["monthlyMobileQcCnt"] == "< 10":
            continue
        elif float(keyword["monthlyMobileQcCnt"]) >= 1000:
            api_result.append(keyword)

    return api_result


# 네이버 검색 API 사용을 통한 총 상품수와 카테고리 추출
def get_products_info(coreKeyword):
    url = "https://openapi.naver.com/v1/search/shop.json?query=" + coreKeyword
    headers = {
        "X-Naver-Client-Id": naver_client_id,
        "X-Naver-Client-Secret": naver_client_secret,
    }
    response = requests.get(url, headers=headers).json()
    total = response["total"]
    try:
        first_item = response["items"][0]
        category = first_item["category1"]
        for i in range(2, 5):
            temp_category = first_item["category" + str(i)]
            if temp_category != "":
                category = category + " > " + temp_category
    except:
        category = "-"

    return {"total": total, "category": category}


# 엑셀에서 키워드 리스트 불러오기
def import_temp_keyword_list(coreKeyword):
    load_wb = load_workbook("./" + coreKeyword + "_temp.xlsx", data_only=True)
    load_ws = load_wb["Sheet"]
    temp_keyword_list = []
    for keyword in load_ws["A"]:
        temp_keyword_list.append(keyword.value)
    return temp_keyword_list


# 최종 결과를 엑셀로 저장
def export_result(coreKeyword):
    write_wb = Workbook()
    write_ws = write_wb.create_sheet("Sheet1")
    write_ws = write_wb.active
    write_ws.append(
        [
            "키워드",
            "PC 검색량",
            "모바일 검색량",
            "총 검색량",
            "상품수",
            "비율",
            "PC 클릭률",
            "모바일 클릭률",
            "카테고리",
        ]
    )
    for keyword_info in result["keywords"]:
        write_ws.append(
            [
                keyword_info["키워드"],
                keyword_info["PC 검색량"],
                keyword_info["모바일 검색량"],
                keyword_info["총 검색량"],
                keyword_info["상품수"],
                keyword_info["비율"],
                keyword_info["PC 클릭률"],
                keyword_info["모바일 클릭률"],
                keyword_info["카테고리"],
            ]
        )
    write_wb.save("./" + coreKeyword + ".xlsx")


# 전체 키워드 추출
def get_result(coreKeyword):
    temp_keyword_list = import_temp_keyword_list(coreKeyword)
    for i in range(0, len(temp_keyword_list) // 5 + 1):
        keywords = temp_keyword_list[5 * i : min(5 * i + 5, len(temp_keyword_list))]
        try:
            api_result = naver_searchad_api(keywords)
        except:
            continue
        time.sleep(0.15)

        for keyword_info in api_result:
            if not keyword_info["relKeyword"] in duplicate_check_list:
                duplicate_check_list.append(keyword_info["relKeyword"])
                keyword = keyword_info["relKeyword"]
                products_info = get_products_info(keyword)
                if keyword_info["monthlyPcQcCnt"] == "< 10":
                    pcQcCnt = 10
                else:
                    pcQcCnt = int(float(keyword_info["monthlyPcQcCnt"]))
                if keyword_info["monthlyMobileQcCnt"] == "< 10":
                    mobileQcCnt = 10
                else:
                    mobileQcCnt = int(float(keyword_info["monthlyMobileQcCnt"]))
                totalQcCnt = pcQcCnt + mobileQcCnt
                productsCnt = int(float(products_info["total"]))
                pcCtr = keyword_info["monthlyAvePcCtr"]
                mobileCtr = keyword_info["monthlyAveMobileCtr"]
                category = products_info["category"]
                print(keyword + " ", end="")
                print(productsCnt)

                keyword_instance = {
                    "키워드": keyword,
                    "PC 검색량": pcQcCnt,
                    "모바일 검색량": mobileQcCnt,
                    "총 검색량": totalQcCnt,
                    "상품수": productsCnt,
                    "비율": round(productsCnt / totalQcCnt, 5),
                    "PC 클릭률": pcCtr,
                    "모바일 클릭률": mobileCtr,
                    "카테고리": category,
                }
                result["keywords"].append(keyword_instance)
                time.sleep(0.20)
            else:
                print("skip")


while True:
    result = {"keywords": []}
    duplicate_check_list = []
    search_keyword = input("메인 키워드 입력: ")
    get_result(search_keyword)
    export_result(search_keyword)
